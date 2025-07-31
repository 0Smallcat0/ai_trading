# -*- coding: utf-8 -*-
"""Excel 數據導出器

此模組提供統一的 Excel 數據導出功能，支援自定義報表模板、
批量導出和歷史數據查詢。

主要功能：
- Excel 格式數據導出
- 自定義報表模板
- 多工作表管理
- 數據格式化和樣式設定
- 批量導出功能
- 歷史數據查詢和導出

支援的導出格式：
- 標準 Excel (.xlsx)
- CSV 格式 (.csv)
- 自定義模板導出
- 多工作表組合導出

Example:
    >>> from src.utils.excel_exporter import ExcelExporter
    >>> exporter = ExcelExporter()
    >>> 
    >>> # 導出單個數據表
    >>> exporter.export_dataframe(data, 'output.xlsx', sheet_name='數據')
    >>> 
    >>> # 使用模板導出
    >>> exporter.export_with_template(data, 'template.xlsx', 'output.xlsx')
    >>> 
    >>> # 批量導出
    >>> exporter.batch_export(data_dict, 'batch_output.xlsx')
"""

import logging
import os
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union
import pandas as pd
import numpy as np
from pathlib import Path

# 設定日誌
logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel 數據導出器
    
    提供全面的 Excel 數據導出和報表生成功能。
    
    Attributes:
        config: 導出配置
        
    Example:
        >>> exporter = ExcelExporter({
        ...     'default_template': 'templates/default.xlsx',
        ...     'output_dir': 'exports/',
        ...     'auto_format': True
        ... })
        >>> exporter.export_dataframe(data, 'report.xlsx')
    """
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """初始化 Excel 導出器
        
        Args:
            config: 配置參數
                - default_template: 預設模板路徑
                - output_dir: 輸出目錄
                - auto_format: 是否自動格式化
                - date_format: 日期格式
        """
        self.config = config or {}
        
        # 基本配置
        self.default_template = self.config.get('default_template', None)
        self.output_dir = Path(self.config.get('output_dir', 'exports'))
        self.auto_format = self.config.get('auto_format', True)
        self.date_format = self.config.get('date_format', '%Y-%m-%d')
        
        # 創建輸出目錄
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # 樣式配置
        self.default_styles = {
            'header': {
                'font': {'bold': True, 'color': 'FFFFFF'},
                'fill': {'fill_type': 'solid', 'start_color': '366092'},
                'alignment': {'horizontal': 'center', 'vertical': 'center'},
                'border': {'style': 'thin'}
            },
            'data': {
                'alignment': {'horizontal': 'left', 'vertical': 'center'},
                'border': {'style': 'thin'}
            },
            'number': {
                'number_format': '#,##0.00'
            },
            'percentage': {
                'number_format': '0.00%'
            },
            'date': {
                'number_format': 'yyyy-mm-dd'
            }
        }
        
        logger.info("Excel 導出器初始化完成")
    
    def export_dataframe(self,
                        data: pd.DataFrame,
                        filename: str,
                        sheet_name: str = 'Sheet1',
                        **kwargs) -> str:
        """導出 DataFrame 到 Excel
        
        Args:
            data: 要導出的數據
            filename: 文件名
            sheet_name: 工作表名稱
            **kwargs: 其他參數
            
        Returns:
            導出文件的完整路徑
        """
        try:
            # 確定輸出路徑
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            output_path = self.output_dir / filename
            
            # 數據預處理
            processed_data = self._preprocess_data(data)
            
            # 創建 Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 寫入數據
                processed_data.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=kwargs.get('include_index', False),
                    startrow=kwargs.get('startrow', 0),
                    startcol=kwargs.get('startcol', 0)
                )
                
                # 應用格式化
                if self.auto_format:
                    self._apply_formatting(writer, sheet_name, processed_data, **kwargs)
            
            logger.info(f"數據導出完成: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"數據導出失敗: {e}")
            raise RuntimeError(f"數據導出失敗: {e}") from e
    
    def export_multiple_sheets(self,
                              data_dict: Dict[str, pd.DataFrame],
                              filename: str,
                              **kwargs) -> str:
        """導出多個工作表到一個 Excel 文件
        
        Args:
            data_dict: 工作表名稱到數據的映射
            filename: 文件名
            **kwargs: 其他參數
            
        Returns:
            導出文件的完整路徑
        """
        try:
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            output_path = self.output_dir / filename
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, data in data_dict.items():
                    # 數據預處理
                    processed_data = self._preprocess_data(data)
                    
                    # 寫入數據
                    processed_data.to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=kwargs.get('include_index', False)
                    )
                    
                    # 應用格式化
                    if self.auto_format:
                        self._apply_formatting(writer, sheet_name, processed_data, **kwargs)
            
            logger.info(f"多工作表導出完成: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"多工作表導出失敗: {e}")
            raise RuntimeError(f"多工作表導出失敗: {e}") from e
    
    def export_with_template(self,
                           data: pd.DataFrame,
                           template_path: str,
                           output_filename: str,
                           data_start_row: int = 2,
                           data_start_col: int = 1) -> str:
        """使用模板導出數據
        
        Args:
            data: 要導出的數據
            template_path: 模板文件路徑
            output_filename: 輸出文件名
            data_start_row: 數據開始行（1-based）
            data_start_col: 數據開始列（1-based）
            
        Returns:
            導出文件的完整路徑
        """
        try:
            from openpyxl import load_workbook
            
            # 載入模板
            template_path = Path(template_path)
            if not template_path.exists():
                raise FileNotFoundError(f"模板文件不存在: {template_path}")
            
            workbook = load_workbook(template_path)
            worksheet = workbook.active
            
            # 數據預處理
            processed_data = self._preprocess_data(data)
            
            # 寫入列標題
            for col_idx, column_name in enumerate(processed_data.columns, start=data_start_col):
                worksheet.cell(row=data_start_row - 1, column=col_idx, value=column_name)
            
            # 寫入數據
            for row_idx, (_, row) in enumerate(processed_data.iterrows(), start=data_start_row):
                for col_idx, value in enumerate(row, start=data_start_col):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # 保存文件
            if not output_filename.endswith('.xlsx'):
                output_filename += '.xlsx'
            
            output_path = self.output_dir / output_filename
            workbook.save(output_path)
            
            logger.info(f"模板導出完成: {output_path}")
            return str(output_path)
            
        except ImportError:
            logger.error("缺少 openpyxl 庫，無法使用模板功能")
            raise ImportError("請安裝 openpyxl: pip install openpyxl")
        except Exception as e:
            logger.error(f"模板導出失敗: {e}")
            raise RuntimeError(f"模板導出失敗: {e}") from e
    
    def batch_export(self,
                    data_sources: Dict[str, Any],
                    base_filename: str,
                    export_format: str = 'xlsx') -> List[str]:
        """批量導出數據
        
        Args:
            data_sources: 數據源字典
            base_filename: 基礎文件名
            export_format: 導出格式
            
        Returns:
            導出文件路徑列表
        """
        try:
            exported_files = []
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            for source_name, data in data_sources.items():
                if isinstance(data, pd.DataFrame) and not data.empty:
                    # 生成文件名
                    filename = f"{base_filename}_{source_name}_{timestamp}.{export_format}"
                    
                    if export_format.lower() == 'xlsx':
                        file_path = self.export_dataframe(data, filename, sheet_name=source_name)
                    elif export_format.lower() == 'csv':
                        file_path = self._export_csv(data, filename)
                    else:
                        logger.warning(f"不支援的導出格式: {export_format}")
                        continue
                    
                    exported_files.append(file_path)
            
            logger.info(f"批量導出完成，共 {len(exported_files)} 個文件")
            return exported_files
            
        except Exception as e:
            logger.error(f"批量導出失敗: {e}")
            raise RuntimeError(f"批量導出失敗: {e}") from e
    
    def export_market_watch_data(self,
                                market_data: Dict[str, pd.DataFrame],
                                filename: Optional[str] = None) -> str:
        """導出市場看盤數據
        
        Args:
            market_data: 市場數據字典
            filename: 文件名
            
        Returns:
            導出文件路徑
        """
        try:
            if filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"市場看盤數據_{timestamp}.xlsx"
            
            # 準備工作表數據
            sheets_data = {}
            
            # 自選股數據
            if 'custom_stocks' in market_data:
                sheets_data['自選股'] = market_data['custom_stocks']
            
            # 概念板塊數據
            if 'concept_ranking' in market_data:
                sheets_data['概念板塊'] = market_data['concept_ranking']
            
            # 龍虎榜數據
            if 'billboard' in market_data:
                sheets_data['龍虎榜'] = market_data['billboard']
            
            # 漲停板數據
            if 'limit_up' in market_data:
                sheets_data['漲停板'] = market_data['limit_up']
            
            # 導出多工作表文件
            output_path = self.export_multiple_sheets(sheets_data, filename)
            
            # 添加摘要工作表
            self._add_summary_sheet(output_path, market_data)
            
            return output_path
            
        except Exception as e:
            logger.error(f"市場看盤數據導出失敗: {e}")
            raise RuntimeError(f"市場看盤數據導出失敗: {e}") from e
    
    def _preprocess_data(self, data: pd.DataFrame) -> pd.DataFrame:
        """數據預處理"""
        processed_data = data.copy()
        
        # 處理無限值
        processed_data = processed_data.replace([np.inf, -np.inf], np.nan)
        
        # 處理日期時間列
        for col in processed_data.columns:
            if processed_data[col].dtype == 'datetime64[ns]':
                processed_data[col] = processed_data[col].dt.strftime(self.date_format)
        
        # 處理數值列的精度
        for col in processed_data.select_dtypes(include=[np.number]).columns:
            if processed_data[col].dtype == 'float64':
                processed_data[col] = processed_data[col].round(4)
        
        return processed_data
    
    def _apply_formatting(self, writer, sheet_name: str, data: pd.DataFrame, **kwargs):
        """應用 Excel 格式化"""
        try:
            from openpyxl.styles import Font, Fill, Alignment, Border, Side, PatternFill
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # 設定列寬
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 格式化標題行
            if data.shape[0] > 0:
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for col in range(1, len(data.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
            
            # 添加邊框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
        except ImportError:
            logger.warning("缺少 openpyxl 庫，跳過格式化")
        except Exception as e:
            logger.warning(f"格式化失敗: {e}")
    
    def _export_csv(self, data: pd.DataFrame, filename: str) -> str:
        """導出 CSV 格式"""
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        output_path = self.output_dir / filename
        processed_data = self._preprocess_data(data)
        
        processed_data.to_csv(output_path, index=False, encoding='utf-8-sig')
        
        logger.info(f"CSV 導出完成: {output_path}")
        return str(output_path)
    
    def _add_summary_sheet(self, file_path: str, market_data: Dict[str, pd.DataFrame]):
        """添加摘要工作表"""
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(file_path)
            
            # 創建摘要工作表
            summary_sheet = workbook.create_sheet('摘要', 0)
            
            # 添加標題
            summary_sheet['A1'] = '市場看盤數據摘要'
            summary_sheet['A2'] = f'生成時間: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            
            row = 4
            
            # 統計各工作表數據
            for sheet_name, data in market_data.items():
                if isinstance(data, pd.DataFrame) and not data.empty:
                    summary_sheet[f'A{row}'] = f'{sheet_name}:'
                    summary_sheet[f'B{row}'] = f'{len(data)} 條記錄'
                    row += 1
            
            # 保存文件
            workbook.save(file_path)
            
        except ImportError:
            logger.warning("缺少 openpyxl 庫，跳過摘要工作表")
        except Exception as e:
            logger.warning(f"添加摘要工作表失敗: {e}")
    
    def get_exporter_info(self) -> Dict[str, Any]:
        """獲取導出器資訊
        
        Returns:
            導出器詳細資訊
        """
        return {
            'exporter_name': 'ExcelExporter',
            'version': '1.0.0',
            'config': self.config,
            'output_dir': str(self.output_dir),
            'supported_formats': ['xlsx', 'csv'],
            'features': [
                'dataframe_export',
                'multiple_sheets',
                'template_export',
                'batch_export',
                'auto_formatting',
                'market_watch_export'
            ]
        }
